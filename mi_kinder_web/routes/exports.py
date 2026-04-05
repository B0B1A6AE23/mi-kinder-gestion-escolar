"""Rutas de exportacion de datos."""
import os
import io
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user

from mi_kinder_web.app import get_db
from mi_kinder_web.config import EXPORTS_DIR

exports_bp = Blueprint("exports", __name__)


@exports_bp.route("/exports/students-excel")
@login_required
def students_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    db = get_db()
    active_year = db.execute(
        "SELECT * FROM school_years WHERE is_active = 1"
    ).fetchone()
    year_id = active_year["id"] if active_year else 0

    group_id = request.args.get("group_id", type=int)

    if group_id:
        students = db.execute(
            """SELECT s.*, g.name as group_name
               FROM students s
               JOIN groups_ g ON g.id = s.group_id
               WHERE s.group_id = ? AND s.is_active = 1
               ORDER BY s.last_name, s.first_name""",
            (group_id,),
        ).fetchall()
    else:
        students = db.execute(
            """SELECT s.*, g.name as group_name
               FROM students s
               JOIN groups_ g ON g.id = s.group_id
               WHERE g.school_year_id = ? AND s.is_active = 1
               ORDER BY g.name, s.last_name, s.first_name""",
            (year_id,),
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Alumnos"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Grupo", "Apellido Paterno", "Apellido Materno", "Nombre",
        "CURP", "Fecha Nacimiento", "Género", "Tutor",
        "Teléfono", "Email", "Fecha Inscripción",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, s in enumerate(students, 2):
        gender_label = {"M": "Masculino", "F": "Femenino"}.get(s["gender"] or "", "")
        data = [
            s["group_name"], s["last_name"], s["second_last_name"] or "",
            s["first_name"], s["curp"] or "", s["birth_date"] or "",
            gender_label, s["guardian_name"] or "",
            s["guardian_phone"] or "", s["guardian_email"] or "",
            s["enrollment_date"] or "",
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(students) + 2)
        )
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 3, 35)

    filename = f"alumnos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    output_path = os.path.join(EXPORTS_DIR, filename)
    wb.save(output_path)

    return send_file(output_path, as_attachment=True, download_name=filename)


@exports_bp.route("/exports/attendance-excel")
@login_required
def attendance_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    db = get_db()
    group_id = request.args.get("group_id", type=int)
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")

    if not group_id or not start_date or not end_date:
        flash("Selecciona grupo y fechas.", "error")
        return redirect(url_for("attendance.report"))

    group = db.execute("SELECT * FROM groups_ WHERE id = ?", (group_id,)).fetchone()
    import re
    raw_name = group["name"] if group else "grupo"
    group_name = re.sub(r'_+', '_', re.sub(r'[\\/:*?"<>|\s]+', '_', raw_name)).strip('_')

    students = db.execute(
        """SELECT s.id, s.first_name || ' ' || s.last_name as full_name
           FROM students s
           WHERE s.group_id = ? AND s.is_active = 1
           ORDER BY s.last_name, s.first_name""",
        (group_id,),
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Alumno", "Presentes", "Faltas", "Retardos", "Justificadas", "Total", "% Asistencia"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, student in enumerate(students, 2):
        row = db.execute(
            """SELECT
                 COUNT(CASE WHEN status='present' THEN 1 END) as present,
                 COUNT(CASE WHEN status='absent' THEN 1 END) as absent,
                 COUNT(CASE WHEN status='late' THEN 1 END) as late,
                 COUNT(CASE WHEN status='justified' THEN 1 END) as justified,
                 COUNT(*) as total
               FROM attendance_records
               WHERE student_id = ? AND date BETWEEN ? AND ?""",
            (student["id"], start_date, end_date),
        ).fetchone()

        rate = round(row["present"] * 100 / max(row["total"], 1), 1)
        data = [student["full_name"], row["present"], row["absent"],
                row["late"], row["justified"], row["total"], rate]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    filename = f"asistencia_{group_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = os.path.join(EXPORTS_DIR, filename)
    wb.save(output_path)

    return send_file(output_path, as_attachment=True, download_name=filename)
